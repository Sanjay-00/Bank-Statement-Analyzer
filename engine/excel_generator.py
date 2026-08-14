"""
excel_generator.py - formatted Excel output.

Palette matches the Streamlit app's theme (.streamlit/config.toml) so the
in-app view and the downloaded report read as the same product, not two
different tools bolted together. Styling helpers and the amber "Check
Statement" sentinel-cell convention are carried over from
CIBIL_EXCEL/excel_generator.py's "Check CIBIL" pattern - an unreconciled row
must look visibly different from a clean one, not just say so in a status
column an analyst might not read.
"""

import io
import re
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRAND       = "1E3A8A"   # deep blue - header bands, matches app primaryColor family
BRAND_LIGHT = "EFF4FC"   # tint of BRAND for subtle section backgrounds
WHITE       = "FFFFFF"
LIGHT_GREY  = "F3F6FB"
CHECK_BG    = "FEF3C7"   # amber - "Check Statement"
CHECK_FONT  = "92610A"
VERIFIED_FONT = "15803D"
FAILED_BG   = "FEE2E2"
FAILED_FONT = "B91C1C"
BORDER_CLR  = "D9E1EE"

# Backwards-compatible alias - most of this module was written against the
# name NAVY; kept as one alias rather than touching every call site.
NAVY = BRAND


def _b():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _f(size=10, bold=False, color="000000", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _a(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _banner(ws, span: str, title: str, subtitle: str = None, height: int = 34) -> int:
    """Full-width colour band as a sheet's title strip - the "cover page"
    touch that makes a report look like a finished product rather than a
    raw data dump. Returns the next free row."""
    start_col, end_col = span.split(":")
    ws.merge_cells(f"{start_col}1:{end_col}1")
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _f(size=15, bold=True, color=WHITE)
    cell.fill = _fill(BRAND)
    cell.alignment = _a(v="center")
    ws.row_dimensions[1].height = height
    if subtitle:
        ws.merge_cells(f"{start_col}2:{end_col}2")
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = _f(size=9, italic=True, color="64748B")
        sub.alignment = _a(v="center")
        ws.row_dimensions[2].height = 16
        return 4
    return 3


TXN_COLUMNS = [
    ("Date", 12),
    ("Narration", 55),
    ("Chq/Ref No", 18),
    ("Debit (Rs.)", 16),
    ("Credit (Rs.)", 16),
    ("Balance (Rs.)", 16),
    ("Category", 16),
    ("Status", 14),
]

_CATEGORY_FONT = {
    "salary": VERIFIED_FONT,
    "emi": FAILED_FONT,
    "rent": "7C3AED",
    "bounce": FAILED_FONT,
    "upi": "2563EB",
    "cash": CHECK_FONT,
    "charges": "64748B",
    "interest": "64748B",
    "uncategorized": "94A3B8",
}

_CHECK = "Check Statement"


def generate_excel(result) -> bytes:
    """`result` is an engine.statement.AnalysisResult."""
    wb = Workbook()

    _build_summary_sheet(wb, result)
    _build_redflags_sheet(wb, result)
    _build_fraud_sheet(wb, result)
    _build_due_date_sheet(wb, result)
    _build_monthly_sheet(wb, result)
    _build_category_sheet(wb, result)
    _build_transactions_sheet(wb, result)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_summary_sheet(wb, result) -> None:
    ws = wb.active
    ws.title = "Statement Summary"
    ws.sheet_properties.tabColor = BRAND
    for col, width in zip("ABCDE", (30, 22, 22, 22, 22)):
        ws.column_dimensions[col].width = width

    generated = datetime.datetime.now().strftime("%d %b %Y, %H:%M")
    title = f"{result.account_holder}  ·  {result.bank_name}" if result.account_holder else result.bank_name
    next_row = _banner(
        ws, "A:E", f"{title}  ·  Bank Statement Analysis",
        subtitle=f"Generated {generated}",
    )

    s = result.summary
    rows = [
        ("Pages Processed", result.page_count),
        ("Scanned Pages (unsupported in v1)", result.scanned_pages),
        ("Opening Balance (Rs.)", s.get("opening_balance")),
        ("Closing Balance (Rs.)", s.get("closing_balance")),
        ("Total Debit (Rs.)", s.get("total_debit")),
        ("Total Credit (Rs.)", s.get("total_credit")),
        ("Transactions Parsed", s.get("transaction_count")),
        ("Verified Rows", s["counts"].get("VERIFIED", 0)),
        ("Failed Reconciliation", s["counts"].get("FAILED", 0)),
        ("Unverified (Check Statement)", s["counts"].get("UNVERIFIED", 0)),
        ("Fraud/Tamper Signals (see Fraud Flags sheet)", len(result.fraud_signals)),
    ]

    r = next_row
    summary_start = r
    for label, value in rows:
        row_bg = _fill(LIGHT_GREY) if (r - summary_start) % 2 == 1 else _fill(WHITE)
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = _f(bold=True, color=BRAND)
        lbl.fill = row_bg
        val = ws.cell(row=r, column=2, value=value if value is not None else "NA")
        val.fill = row_bg
        if isinstance(value, (int, float)):
            val.number_format = "#,##0.00"
        if label == "Failed Reconciliation" and value:
            val.fill = _fill(FAILED_BG)
            val.font = _f(bold=True, color=FAILED_FONT)
        elif label == "Unverified (Check Statement)" and value:
            val.fill = _fill(CHECK_BG)
            val.font = _f(bold=True, color=CHECK_FONT)
        elif label.startswith("Fraud/Tamper Signals") and value:
            val.fill = _fill(FAILED_BG)
            val.font = _f(bold=True, color=FAILED_FONT)
        r += 1

    ws.freeze_panes = f"A{next_row}"


_SEVERITY_STYLE = {
    "HIGH":   (FAILED_BG, FAILED_FONT),
    "MEDIUM": (CHECK_BG, CHECK_FONT),
    "LOW":    (LIGHT_GREY, "595959"),
}


def _kv_pair(ws, row, col, label, value, number_format=None):
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = _f(bold=True, color=BRAND)
    val = ws.cell(row=row, column=col + 1, value=value if value is not None else "NA")
    if number_format and isinstance(value, (int, float)):
        val.number_format = number_format
    return row + 1


def _build_redflags_sheet(wb, result) -> None:
    """
    The underwriting-signals view: ABB, salary consistency, bounce
    frequency, cash-flow ratios, then the red flags those signals feed into
    (engine/signals/redflags.py - each flag reads a number from one of the
    sections above it and names the number that triggered it).
    """
    ws = wb.create_sheet("Red Flags")
    ws.sheet_properties.tabColor = FAILED_FONT
    for col, width in zip("ABC", (30, 24, 90)):
        ws.column_dimensions[col].width = width

    next_row = _banner(
        ws, "A:C", "Red flags & underwriting signals",
        subtitle="ABB, salary consistency, bounce frequency, cash-flow ratios, and the flags derived from them.",
    )
    r = next_row

    sub = ws.cell(row=r, column=1, value="Cash-flow score")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    score = result.score
    score_cell = ws.cell(row=r, column=2, value=f"{score['score']:.0f} / 1000")
    score_cell.font = _f(size=14, bold=True, color=BRAND)
    ws.cell(row=r, column=1, value="Composite score").font = _f(bold=True, color=BRAND)
    r += 1
    r = _kv_pair(ws, r, 1, "Volatility (0=stable, 1=volatile)", f"{score['volatility']:.2f}")
    r = _kv_pair(ws, r, 1, "Red-flag penalty applied", f"-{score['red_flag_penalty']}")
    comp_note = ws.cell(row=r, column=1,
                         value="Components (0-100 each, weighted): " +
                               ", ".join(f"{k}={v} (w={score['weights'][k]})" for k, v in score["components"].items()))
    comp_note.font = _f(size=8, italic=True, color="64748B")
    ws.merge_cells(f"A{r}:C{r}")
    comp_note.alignment = _a("left", wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1
    note = ws.cell(row=r, column=1,
                    value="Policy-default weights, not a fitted/validated model - tune before relying on this for real decisions.")
    note.font = _f(size=8, italic=True, color="94A3B8")
    ws.merge_cells(f"A{r}:C{r}")
    r += 2

    sub = ws.cell(row=r, column=1, value="FOIR / DSCR")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    fd = score["foir_dscr"]
    r = _kv_pair(ws, r, 1, "Estimated monthly income",
                  f"Rs.{fd['monthly_income']:,.2f} ({fd['income_source']})" if fd["monthly_income"] else "NA")
    r = _kv_pair(ws, r, 1, "Monthly fixed obligations (EMI+rent)", f"Rs.{fd['monthly_fixed_obligations']:,.2f}")
    r = _kv_pair(ws, r, 1, "FOIR", f"{fd['foir']:.1f}% - {fd['foir_band']}" if fd["foir"] is not None else "NA")
    r = _kv_pair(ws, r, 1, "DSCR", f"{fd['dscr']:.2f} - {fd['dscr_band']}" if fd["dscr"] is not None else "NA (no EMI debt service detected)")
    r += 1

    sub = ws.cell(row=r, column=1, value="Average Bank Balance (ABB)")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    abb = result.abb
    for w in (1, 3, 6, 12):
        win = abb["windows"].get(w, {})
        avg = win.get("average")
        label = f"ABB{w} (last {w} month{'s' if w > 1 else ''})"
        detail = f"{avg:,.2f}" if avg is not None else f"NA ({win.get('covered_days', 0)}/{win.get('total_days', 0)} days covered)"
        r = _kv_pair(ws, r, 1, label, detail)
    r += 1

    sub = ws.cell(row=r, column=1, value="Salary consistency")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    sal = result.salary
    r = _kv_pair(ws, r, 1, "Salary detected", "Yes" if sal["detected"] else "No")
    if sal["detected"]:
        r = _kv_pair(ws, r, 1, "Recurrence",
                      f"{sal['months_with_salary']} of {sal['months_covered']} recent month(s) ({sal['recurrence_rate']:.0%})")
        r = _kv_pair(ws, r, 1, "Average amount",
                      f"Rs.{sal['average_amount']:,.2f}" if sal["average_amount"] else "NA")
        r = _kv_pair(ws, r, 1, "Amount variation",
                      f"{sal['coefficient_of_variation']:.0%}" if sal["coefficient_of_variation"] is not None else "NA")
        r = _kv_pair(ws, r, 1, "Irregular", "Yes" if sal["irregular"] else "No")
    r += 1

    sub = ws.cell(row=r, column=1, value="Bounce / return frequency")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    bnc = result.bounces
    r = _kv_pair(ws, r, 1, "Bounce/return count", bnc["count"])
    r = _kv_pair(ws, r, 1, "Rate per month",
                  f"{bnc['rate_per_month']:.2f}" if bnc["rate_per_month"] is not None else "NA")
    r += 1

    sub = ws.cell(row=r, column=1, value="Cash-flow ratios")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    cf = result.cashflow
    cdr = cf.get("cash_dependency_ratio")
    r = _kv_pair(ws, r, 1, "Cash dependency", f"{cdr:.0%}" if cdr is not None else "NA")
    ec = cf.get("expense_concentration")
    r = _kv_pair(ws, r, 1, "Expense concentration",
                  f"{ec['ratio']:.0%} in \"{ec['category'].title()}\"" if ec else "NA")
    r += 2

    sub = ws.cell(row=r, column=1, value="Red flags")
    sub.font = _f(size=13, bold=True, color=BRAND)
    r += 1

    flags = result.red_flags
    if not flags:
        c = ws.cell(row=r, column=1, value="No red flags raised.")
        c.font = _f(italic=True, color=VERIFIED_FONT)
        ws.freeze_panes = f"A{next_row}"
        return

    headers = ["Severity", "Signal", "Detail"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
    r += 1

    for flag in flags:
        bg, fg = _SEVERITY_STYLE.get(flag.get("severity"), (LIGHT_GREY, "000000"))
        values = [flag.get("severity", ""), flag.get("code", ""), flag.get("detail", "")]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.border = _b()
            c.fill = _fill(bg)
            c.alignment = _a("left", wrap=(col_idx == 3))
            if col_idx == 1:
                c.font = _f(bold=True, color=fg)
            else:
                c.font = _f(color=fg if col_idx == 2 else "000000")
        ws.row_dimensions[r].height = 30
        r += 1

        instances = flag.get("instances")
        if instances:
            r = _write_instance_table(ws, r, instances)

    ws.freeze_panes = f"A{next_row}"


def _build_fraud_sheet(wb, result) -> None:
    """
    Cheap, deterministic tamper/anomaly signals (PDF metadata, structural
    edit-history, font consistency, round-tripping, duplicate rows, round-
    number clustering) - see engine/extract/pdf_forensics.py and
    engine/signals/fraud.py for what each check actually does and why.
    Every row here is a signal for an analyst to look at, not a verdict.
    """
    ws = wb.create_sheet("Fraud Flags")
    ws.sheet_properties.tabColor = FAILED_FONT
    for col, width in zip("ABC", (14, 26, 90)):
        ws.column_dimensions[col].width = width

    signals = result.fraud_signals
    next_row = _banner(
        ws, "A:C", "Fraud / Tamper Signals",
        subtitle="Deterministic checks only, each row is a signal to review, not a verdict.",
    )

    if not signals:
        c = ws.cell(row=next_row, column=1, value="No fraud/tamper signals detected.")
        c.font = _f(italic=True, color=VERIFIED_FONT)
        return

    headers = ["Severity", "Signal", "Detail"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=next_row, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()

    row_num = next_row + 1
    for sig in signals:
        bg, fg = _SEVERITY_STYLE.get(sig.get("severity"), (LIGHT_GREY, "000000"))
        values = [sig.get("severity", ""), sig.get("code", ""), sig.get("detail", "")]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            c.border = _b()
            c.fill = _fill(bg)
            c.alignment = _a("left", wrap=(col_idx == 3))
            if col_idx == 1:
                c.font = _f(bold=True, color=fg)
            else:
                c.font = _f(color=fg if col_idx == 2 else "000000")
        ws.row_dimensions[row_num].height = 30
        row_num += 1

        instances = sig.get("instances")
        if instances:
            row_num = _write_instance_table(ws, row_num, instances)

    ws.freeze_panes = f"A{next_row + 1}"


def _write_instance_table(ws, row_num: int, instances: list) -> int:
    """Compact indented sub-table of a fraud signal's individual instances
    (round-trip pairs, duplicate rows, ...) directly under its summary row -
    the tabular alternative to jamming every instance into one run-on
    sentence in the Detail cell."""
    cols = list(instances[0].keys())
    for col_idx, h in enumerate(cols, 2):  # start at column B, indented under Detail
        c = ws.cell(row=row_num, column=col_idx, value=h)
        c.font = _f(size=9, bold=True, color=BRAND)
        c.fill = _fill(BRAND_LIGHT)
        c.border = _b()
        c.alignment = _a("left")
    row_num += 1

    for inst in instances:
        for col_idx, key in enumerate(cols, 2):
            v = inst[key]
            c = ws.cell(row=row_num, column=col_idx, value=v)
            c.border = _b()
            c.font = _f(size=9)
            if isinstance(v, float):
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
            else:
                c.alignment = _a("left")
        row_num += 1

    return row_num + 1  # blank spacer row before the next signal


_MONTH_NAME = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_ANCHOR_HEADERS = [
    (4, "4th (for 5th due)"),
    (9, "9th (for 10th due)"),
    (14, "14th (for 15th due)"),
    (19, "19th (for 20th due)"),
]

_PRIORITY_BG = {1: "DCFCE7", 2: "FEF3C7", 3: "FFEDD5"}
_PRIORITY_FONT = {1: VERIFIED_FONT, 2: CHECK_FONT, 3: "C2410C"}


def _build_due_date_sheet(wb, result) -> None:
    """
    EMI due-date recommendation: closing balance on the day before each of
    the four common due dates (5th/10th/15th/20th), averaged over the last
    N months, ranked so the analyst can propose the due date the customer
    is most likely to have funds ready for.
    """
    ws = wb.create_sheet("Due Date Analysis")
    ws.sheet_properties.tabColor = "2563EB"
    for col, width in zip("ABCDEF", (10, 24, 30, 22, 20, 20)):
        ws.column_dimensions[col].width = width

    dda = result.due_date_analysis
    next_row = _banner(
        ws, "A:D", "Recommended EMI due date",
        subtitle="Ranked by average closing balance on the day before each due date.",
    )

    r = next_row
    headers = ["Priority", "Recommended Due Date", "Anchor Day Checked",
               f"Avg Balance (Rs.), last {dda['months_used']} mo"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
        c.alignment = _a("left" if col_idx in (2, 3) else "center", wrap=True)
    r += 1

    if not dda["recommendations"]:
        c = ws.cell(row=r, column=1, value="Not enough reconciled data to recommend a due date.")
        c.font = _f(italic=True)
        r += 1
    for rec in dda["recommendations"]:
        bg = _PRIORITY_BG.get(rec["priority"])
        fg = _PRIORITY_FONT.get(rec["priority"], "000000")
        values = [rec["priority"], f"{rec['due_date']} of the month",
                  f"{rec['anchor_day']}{'th' if rec['anchor_day'] not in (1,2,3) else ''} (day before due date)",
                  rec["avg_balance"]]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.border = _b()
            if bg:
                c.fill = _fill(bg)
                c.font = _f(bold=True, color=fg)
            if col_idx == 4:
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
            elif col_idx in (2, 3):
                c.alignment = _a("left")
            else:
                c.alignment = _a("center")
        r += 1

    r += 2
    sub = ws.cell(row=r, column=1, value="Month-by-Month Closing Balance on Each Anchor Day")
    sub.font = _f(size=12, bold=True, color=NAVY)
    r += 1

    hdr_row = r
    ws.cell(row=hdr_row, column=1, value="Month").font = _f(bold=True, color=WHITE)
    ws.cell(row=hdr_row, column=1).fill = _fill(NAVY)
    ws.cell(row=hdr_row, column=1).border = _b()
    for col_idx, (_anchor, label) in enumerate(_ANCHOR_HEADERS, 2):
        c = ws.cell(row=hdr_row, column=col_idx, value=label)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.border = _b()
        c.alignment = _a("center", wrap=True)
    r += 1

    for row in dda["monthly"]:
        label = f"{_MONTH_NAME[row['month']]} {row['year']}"
        c = ws.cell(row=r, column=1, value=label)
        c.font = _f(bold=True)
        c.border = _b()
        for col_idx, (anchor, _label) in enumerate(_ANCHOR_HEADERS, 2):
            val = row.get(anchor)
            c2 = ws.cell(row=r, column=col_idx, value=val if val is not None else "NA")
            c2.border = _b()
            if isinstance(val, (int, float)):
                c2.number_format = "#,##0.00"
                c2.alignment = _a("right")
            else:
                c2.alignment = _a("center")
        r += 1

    # Average row - same lookback window as the recommendation cards above,
    # so the two sections of this sheet agree with each other.
    avg_label = ws.cell(row=r, column=1, value=f"Average (last {dda['months_used']} mo)")
    avg_label.font = _f(bold=True, color=WHITE)
    avg_label.fill = _fill(BRAND)
    avg_label.border = _b()
    for col_idx, (anchor, _label) in enumerate(_ANCHOR_HEADERS, 2):
        val = dda["averages"].get(anchor)
        c2 = ws.cell(row=r, column=col_idx, value=val if val is not None else "NA")
        c2.border = _b()
        c2.fill = _fill(BRAND)
        c2.font = _f(bold=True, color=WHITE)
        if isinstance(val, (int, float)):
            c2.number_format = "#,##0.00"
            c2.alignment = _a("right")
        else:
            c2.alignment = _a("center")
    r += 1

    ws.freeze_panes = f"A{next_row}"


def _build_monthly_sheet(wb, result) -> None:
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = "16A34A"
    for col, width in zip("ABCDEFG", (14, 18, 18, 18, 14, 14, 14)):
        ws.column_dimensions[col].width = width

    next_row = _banner(ws, "A:G", "Monthly summary", subtitle="Total debit/credit/net cash flow, month by month.")

    headers = ["Month", "Total Debit (Rs.)", "Total Credit (Rs.)", "Net (Rs.)",
               "Debit Count", "Credit Count", "Txn Count"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=next_row, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
        c.alignment = _a("center", wrap=True)
    ws.row_dimensions[next_row].height = 22

    row_num = next_row + 1
    for (year, month), b in result.monthly.items():
        row_bg = _fill(LIGHT_GREY) if row_num % 2 == 0 else _fill(WHITE)
        values = [f"{_MONTH_NAME[month]} {year}", b["debit"], b["credit"], b["net"],
                  b["debit_count"], b["credit_count"], b["txn_count"]]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            c.border = _b()
            c.fill = row_bg
            if col_idx == 1:
                c.font = _f(bold=True)
                c.alignment = _a("left")
            elif col_idx in (2, 3, 4):
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
                if col_idx == 4 and isinstance(v, (int, float)) and v < 0:
                    c.font = _f(color=FAILED_FONT)
            else:
                c.alignment = _a("center")
        row_num += 1

    ws.freeze_panes = f"A{next_row + 1}"


def _build_category_sheet(wb, result) -> None:
    ws = wb.create_sheet("Category Summary")
    ws.sheet_properties.tabColor = "7C3AED"
    for col, width in zip("ABCDE", (20, 14, 18, 18, 18)):
        ws.column_dimensions[col].width = width

    next_row = _banner(ws, "A:E", "Category summary",
                        subtitle="Rule-based narration matching. Uncategorized means no rule matched, not a guess.")

    headers = ["Category", "Txn Count", "Total Debit (Rs.)", "Total Credit (Rs.)", "Net (Rs.)"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=next_row, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
        c.alignment = _a("center", wrap=True)
    ws.row_dimensions[next_row].height = 22

    row_num = next_row + 1
    for cat, b in result.category_summary.items():
        row_bg = _fill(LIGHT_GREY) if row_num % 2 == 0 else _fill(WHITE)
        net = b["total_credit"] - b["total_debit"]
        label = cat.replace("_", " ").title()
        values = [label, b["count"], b["total_debit"], b["total_credit"], net]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            c.border = _b()
            c.fill = row_bg
            if col_idx == 1:
                c.font = _f(bold=True, color=_CATEGORY_FONT.get(cat, "94A3B8"))
                c.alignment = _a("left")
            elif col_idx == 2:
                c.alignment = _a("center")
            else:
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
        row_num += 1

    ws.freeze_panes = f"A{next_row + 1}"


def _build_transactions_sheet(wb, result) -> None:
    ws = wb.create_sheet("Transactions")
    ws.sheet_properties.tabColor = "64748B"
    for col_idx, (_, width) in enumerate(TXN_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    next_row = _banner(ws, f"A:{get_column_letter(len(TXN_COLUMNS))}", "Transactions",
                        subtitle="Full reconciled ledger. Amber rows need a manual check against the source PDF.")

    hdr_fill = _fill(BRAND)
    for col_idx, (header, _) in enumerate(TXN_COLUMNS, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=header)
        cell.font = _f(size=11, bold=True, color=WHITE)
        cell.fill = hdr_fill
        cell.alignment = _a("center", wrap=True)
        cell.border = _b()
    ws.row_dimensions[next_row].height = 24

    row_num = next_row + 1
    for t in result.transactions:
        row_bg = _fill(LIGHT_GREY) if row_num % 2 == 0 else _fill(WHITE)

        date_val = t.date if isinstance(t.date, datetime.date) else (t.date or "NA")
        debit_cell = _CHECK if (t.status == "UNVERIFIED" and t.debit is None and t.status != "OPENING") else t.debit
        credit_cell = _CHECK if (t.status == "UNVERIFIED" and t.credit is None and t.status != "OPENING") else t.credit

        category_label = (t.category or "uncategorized").replace("_", " ").title()
        values = [date_val, t.narration, t.chq_ref or "", t.debit, t.credit, t.balance, category_label, t.status]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = _b()
            cell.fill = row_bg

            if col_idx == 1:  # Date
                cell.alignment = _a("center")
                if isinstance(value, datetime.date):
                    cell.number_format = "DD-MMM-YYYY"
            elif col_idx == 2:  # Narration
                cell.alignment = _a("left", wrap=True)
                cell.font = _f(size=9)
            elif col_idx == 3:  # Chq/Ref
                cell.alignment = _a("center")
                cell.font = _f(size=9)
            elif col_idx in (4, 5, 6):  # amount columns
                if value is None:
                    cell.value = ""
                    cell.alignment = _a("center")
                else:
                    cell.number_format = '#,##0.00;-#,##0.00;"-"'
                    cell.alignment = _a("right")
            elif col_idx == 7:  # Category
                cell.alignment = _a("center")
                cell.font = _f(size=9, bold=True, color=_CATEGORY_FONT.get(t.category, "94A3B8"))
            elif col_idx == 8:  # Status
                cell.alignment = _a("center")
                if value == "VERIFIED":
                    cell.font = _f(bold=True, color=VERIFIED_FONT)
                elif value == "FAILED":
                    cell.fill = _fill(FAILED_BG)
                    cell.font = _f(bold=True, color=FAILED_FONT)
                elif value == "UNVERIFIED":
                    cell.fill = _fill(CHECK_BG)
                    cell.font = _f(bold=True, color=CHECK_FONT)
                elif value == "OPENING":
                    cell.font = _f(bold=True, color=NAVY)

        # Amber "Check Statement" flag on the row itself when unverified,
        # same visual convention as CIBIL_EXCEL's "Check CIBIL" cells.
        if t.status == "UNVERIFIED":
            for col_idx in (4, 5):
                c = ws.cell(row=row_num, column=col_idx)
                if c.value in (None, ""):
                    c.value = _CHECK
                    c.fill = _fill(CHECK_BG)
                    c.font = _f(italic=True, bold=True, color=CHECK_FONT)
                    c.alignment = _a("center")

        ws.row_dimensions[row_num].height = 30
        row_num += 1

    ws.freeze_panes = f"A{next_row + 1}"
    ws.print_title_rows = f"1:{next_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


def generate_quick_excel(result) -> bytes:
    """`result` is an engine.statement.QuickAnalysisResult. One sheet, not
    seven: ABB, due-date recommendation, and monthly credit/debit only -
    the same condensed scope as the quick-analysis UI tab, so the download
    and the on-screen view never promise different things."""
    wb = Workbook()
    _build_quick_sheet(wb, result)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_quick_sheet(wb, result) -> None:
    ws = wb.active
    ws.title = "Quick Summary"
    ws.sheet_properties.tabColor = BRAND
    for col, width in zip("ABCD", (30, 24, 24, 24)):
        ws.column_dimensions[col].width = width

    generated = datetime.datetime.now().strftime("%d %b %Y, %H:%M")
    title = f"{result.account_holder}  ·  {result.bank_name}" if result.account_holder else result.bank_name
    next_row = _banner(
        ws, "A:D", f"{title}  ·  Quick Analysis",
        subtitle=f"Generated {generated}  ·  Average balance, best EMI due date, monthly cash flow",
    )
    r = next_row

    s = result.summary
    info_rows = [
        ("Pages Processed", result.page_count),
        ("Transactions Parsed", s.get("transaction_count")),
        ("Opening Balance (Rs.)", s.get("opening_balance")),
        ("Closing Balance (Rs.)", s.get("closing_balance")),
    ]
    for label, value in info_rows:
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = _f(bold=True, color=BRAND)
        val = ws.cell(row=r, column=2, value=value if value is not None else "NA")
        if isinstance(value, (int, float)):
            val.number_format = "#,##0.00"
        r += 1
    r += 1

    sub = ws.cell(row=r, column=1, value="Average Bank Balance (ABB)")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    abb = result.abb
    for w in (1, 3, 6, 12):
        win = abb["windows"].get(w, {})
        avg = win.get("average")
        label = f"ABB{w} (last {w} month{'s' if w > 1 else ''})"
        detail = f"Rs.{avg:,.2f}" if avg is not None else f"NA ({win.get('covered_days', 0)}/{win.get('total_days', 0)} days covered)"
        r = _kv_pair(ws, r, 1, label, detail)
    r += 1

    sub = ws.cell(row=r, column=1, value="Recommended EMI due date")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    dda = result.due_date_analysis
    headers = ["Priority", "Recommended Due Date", f"Avg Balance (Rs.), last {dda['months_used']} mo"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
        c.alignment = _a("left" if col_idx == 2 else "center", wrap=True)
    r += 1

    if not dda["recommendations"]:
        c = ws.cell(row=r, column=1, value="Not enough reconciled data to recommend a due date.")
        c.font = _f(italic=True)
        r += 1
    for rec in dda["recommendations"]:
        bg = _PRIORITY_BG.get(rec["priority"])
        fg = _PRIORITY_FONT.get(rec["priority"], "000000")
        values = [rec["priority"], f"{rec['due_date']} of the month", rec["avg_balance"]]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.border = _b()
            if bg:
                c.fill = _fill(bg)
                c.font = _f(bold=True, color=fg)
            if col_idx == 3:
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
            elif col_idx == 2:
                c.alignment = _a("left")
            else:
                c.alignment = _a("center")
        r += 1
    r += 1

    sub = ws.cell(row=r, column=1, value="Monthly credit / debit")
    sub.font = _f(size=12, bold=True, color=BRAND)
    r += 1
    headers = ["Month", "Total Debit (Rs.)", "Total Credit (Rs.)", "Net (Rs.)"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(BRAND)
        c.border = _b()
        c.alignment = _a("center", wrap=True)
    r += 1

    for (year, month), b in result.monthly.items():
        row_bg = _fill(LIGHT_GREY) if r % 2 == 0 else _fill(WHITE)
        values = [f"{_MONTH_NAME[month]} {year}", b["debit"], b["credit"], b["net"]]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.border = _b()
            c.fill = row_bg
            if col_idx == 1:
                c.font = _f(bold=True)
                c.alignment = _a("left")
            else:
                c.number_format = "#,##0.00"
                c.alignment = _a("right")
                if col_idx == 4 and isinstance(v, (int, float)) and v < 0:
                    c.font = _f(color=FAILED_FONT)
        r += 1

    ws.freeze_panes = f"A{next_row}"


def get_filename(bank_name: str, account_holder: str = None, label: str = "Statement_Analysis") -> str:
    """Named after the customer when the header block yielded a confident
    name; falls back to the bank name otherwise - never a guessed name."""
    label_part = account_holder or bank_name
    safe_name = re.sub(r"[^\w\s-]", "", label_part).strip().replace(" ", "_")
    date_str = datetime.datetime.now().strftime("%d%b%Y")
    return f"{safe_name}_{label}_{date_str}.xlsx"
